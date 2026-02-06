# ------------------------------------------------------------------------------------------------------------------
# File Name:    test_EOL_Parameters_Data_For_MPI_file.py
# Author:       Srinivasa V
# Created:      10 Oct 2024
# Description:  This File contains Test case for reading EOL parameters and creating a excel file report.
# ------------------------------------------------------------------------------------------------------------------
# Copyright (C) 2024 APTIV
#
# This file is part of the MnM HPCC project - Manufacture Diagnostics Component.
#
# Manufacture Diagnostics Project is free software: you can redistribute it
# and/or modify it under the terms of Diagnostics Department ASUX.
# ------------------------------------------------------------------------------------------------------------------
import time
from datetime import datetime

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from infotest.utilities.util_excel_report import adjust_column_row_width_as_per_length
from infotest.core.trace_logger import app_logger, print_owl_symbol
from infotest.utilities.util_man_diag import hex_to_ascii_from_n

index_row = None
wb = None
ws = None
SOC_SoftwareVersion = "XX.YY.ZZ"
timestamp = None

# Define custom IDs corresponding to each test case
custom_ids = (
    "Boot software id",
    "Application software id",
    "ECU software Version",
    "ECU silver box HW version",
    "Data Supplier BSP Version for AG550 (5G)",
    "Data Supplier BSP Version for AG215 (Cv2X)",
    "Data Supplier eUICCID",
    "Data Supplier IMEI",
    "MESN_NUMBER",
    "Data Supplier ICCID",
    "BT Firmware version",
    "SB_CUST_PART_NUM",
    "SYS_SUPPLY_ID",
    "ECU_MAN_DATE",
    "ECU_SERIAL_NUM",
    "SYS_SUPPLIER_PART",
    "VIN",
    "SYS_ENG_TYPE",
    "REPAIR_SHOP_CODE",
    "PROGRAMMING_DATE",
    "ECU_INSTALL_DATE",
    "MODEL_NUMBER",
    "VARIANT_CODE",
    "FEATURE_CODE_IS",
    "FEATURE_CODE_IC",
    "MCM_CV2X_FC",
    "MCM_SDV Feature code",
    "MCM_CONNECT_FC",
    "SITE_CODE",
    "SHIFT_ID",
    "DAILY_SEQ_NUMBER",
    "SB_PART_ID_CODE",
    "SB_SPPL_CODE_NUM",
    "PCBA_PART_NUM",
    "SBM_Reserved",
    "checksum la_super",
    "checksum modem_a",
    "checksum modem_b",
    "checksum bluetooth_a",
    "checksum bluetooth_b",
    "checksum la_vms_a",
    "checksum la_vms_b",
    "checksum sw_info_a",
    "checksum sw_info_b",
    "checksum vip_a",
    "checksum vip_b",
    "checksum ifs2_a",
    "checksum ifs2_b",
    "checksum system_a",
    "checksum system_b",
)

#   (Sl No.,  EOL Parameter,          ExpDataLength, ActDataLength, Data_StartPos  Enter_ManDiag  MFG Command)

TEST_CASES = [
    (1, "Boot software id", "1C", "0x1B", 3, True, "6D643E 64 02 00 00 01 00"),
    (2, "Application software id", "1C", "0x1B", 3, False, "6D643E 64 02 00 00 01 01"),
    (3, "ECU software Version", "79", "0x78", 3, False, "6D643E 64 02 00 00 01 02"),
    (4, "ECU silver box HW version", "05", "0x04", 3, False, "6D643E 78 01 00 00 01 06"),
    (5, "Data Supplier BSP Version for AG550", "21", "0x20", 3, False, "6D643E 6E 13 00 00 01 09"),
    (6, "Data Supplier BSP Version for AG215", "21", "0x20", 3, False, "6D643E 6E 13 00 00 01 0A"),
    (7, "Data Supplier eUICCID", "21", "0x20", 3, False, "6D643E 6E 13 00 00 01 03"),
    (8, "Data Supplier IMEI", "10", "0x0F", 3, False, "6D643E 6E 13 00 00 01 02"),
    (9, "MESN_NUMBER", "0C", "0x0A", 6, False, "6D643E 66 11 00 00 01 18"),
    (10, "Data Supplier ICCID", "15", "0x14", 3, False, "6D643E 6E 13 00 00 01 05"),
    (11, "BT Firmware version", "0D", "0x0C", 3, False, "6D643E 64 01 00 00 01 00"),
    (12, "SB_CUST_PART_NUM", "12", "0x10", 6, False, "6D643E 66 11 00 00 01 0A"),
    (13, "SYS_SUPPLY_ID", "12", "0x10", 6, False, "6D643E 66 11 00 00 01 0B"),
    (14, "ECU_MAN_DATE", "08", "0x06", 6, False, "6D643E 66 11 00 00 01 0C"),
    (15, "ECU_SERIAL_NUM", "0F", "0x0D", 6, False, "6D643E 66 11 00 00 01 0D"),
    (16, "SYS_SUPPLIER_PART", "0F", "0x0D", 6, False, "6D643E 66 11 00 00 01 0E"),
    (17, "VIN", "13", "0x11", 6, False, "6D643E 66 11 00 00 01 0F"),
    (18, "SYS_ENG_TYPE", "12", "0x10", 6, False, "6D643E 66 11 00 00 01 10"),
    (19, "REPAIR_SHOP_CODE", "12", "0x10", 6, False, "6D643E 66 11 00 00 01 11"),
    (20, "PROGRAMMING_DATE", "08", "0x06", 6, False, "6D643E 66 11 00 00 01 12"),
    (21, "ECU_INSTALL_DATE", "08", "0x06", 6, False, "6D643E 66 11 00 00 01 13"),
    (22, "MODEL_NUMBER", "14", "0x12", 6, False, "6D643E 66 11 00 00 01 14"),
    (23, "VARIANT_CODE", "16", "0x14", 6, False, "6D643E 66 11 00 00 01 15"),
    (24, "FEATURE_CODE_IS", "3E", "0x3C", 6, False, "6D643E 66 11 00 00 01 16"),
    (25, "FEATURE_CODE_IC", "3E", "0x3C", 6, False, "6D643E 66 11 00 00 01 17"),
    (26, "MCM_CV2X_FC", "0C", "0x0A", 6, False, "6D643E 66 11 00 00 01 08"),
    (27, "MCM_SDV Feature code", "1D", "0x20", 6, False, "6D643E 66 01 00 00 05 FF 20 31 C4 18"),
    (28, "MCM_CONNECT_FC", "3F", "0x3D", 6, False, "6D643E 66 11 00 00 01 07"),
    (29, "SITE_CODE", "03", "0x01", 6, False, "6D643E 66 11 00 00 01 00"),
    (30, "SHIFT_ID", "03", "0x01", 6, False, "6D643E 66 11 00 00 01 01"),
    (31, "DAILY_SEQ_NUMBER", "06", "0x04", 6, False, "6D643E 66 11 00 00 01 02"),
    (32, "SB_PART_ID_CODE", "05", "0x03", 6, False, "6D643E 66 11 00 00 01 03"),
    (33, "SB_SPPL_CODE_NUM", "09", "0x07", 6, False, "6D643E 66 11 00 00 01 04"),
    (34, "PCBA_PART_NUM", "0A", "0x08", 6, False, "6D643E 66 11 00 00 01 05"),
    (35, "SBM_Reserved", "0E", "0x0C", 6, False, "6D643E 66 11 00 00 01 06"),
    (36, "checksum la_super", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 00"),
    (37, "checksum modem_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 01"),
    (38, "checksum modem_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 02"),
    (39, "checksum bluetooth_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 03"),
    (40, "checksum bluetooth_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 04"),
    (41, "checksum la_vms_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 05"),
    (42, "checksum la_vms_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 06"),
    (43, "checksum sw_info_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 07"),
    (44, "checksum sw_info_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 08"),
    (45, "checksum vip_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 09"),
    (46, "checksum vip_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 0A"),
    (47, "checksum ifs2_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 0B"),
    (48, "checksum ifs2_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 0C"),
    (49, "checksum system_a", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 0D"),
    (50, "checksum system_b", "05", "0x04", 3, False, "6D643E 64 04 00 00 01 0E"),
]


# dependency functions
def Create_excelFile_with_table_headings(ManDiag_component) -> bool:
    global index_row, wb, ws, SOC_SoftwareVersion, timestamp

    index_row = 5
    # Create a workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EOL_PARAMETERS"
    ws['B2'] = "Software Version"
    ws['C2'] = "XX.YY.ZZ"
    ws['A4'] = "Sl No."
    ws['B4'] = "EOL Parameter"
    ws['C4'] = "Data Length (hex)"
    ws['D4'] = "Read Command"
    ws['E4'] = "Data in HEX"

    # Merge cells for the "Software Version" header
    ws["B2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Merge cells for the software version value
    ws["C2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C2"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Pre-condition for BT firmware version read
    actual_test_result = ManDiag_component.send_request_command_and_validate_response("6D643E 3D 01 01 00 01 00")
    assert actual_test_result, f"Failed to Turn OFF BT"
    time.sleep(3)
    actual_test_result = ManDiag_component.send_request_command_and_validate_response("6D643E 3D 01 01 00 01 01")
    assert actual_test_result, f"Failed to Turn ON BT"

    # Read the Sw version and store it in a variable
    response = ManDiag_component.send_request_command_and_get_response("6D643E 6E 12 00 00 01 03")
    if not response:
        app_logger.warning(f"No response for 0x6E12 Sw version")
        ws['C2'] = ILLEGAL_CHARACTERS_RE.sub(r'', "NONE")
    else:
        SOC_SoftwareVersion = hex_to_ascii_from_n(response.get("Databytes", ""), 1)
        ws['C2'] = ILLEGAL_CHARACTERS_RE.sub(r'', SOC_SoftwareVersion)

    # Get the current timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    return True


def apply_tabularFormate_alignment_ToExcelData(diag_test) -> bool:
    global index_row, wb, ws, SOC_SoftwareVersion, timestamp

    # Adjust the alignment
    adjust_column_row_width_as_per_length(ws)

    ws.column_dimensions['E'].width = 100
    for rows in ws.iter_rows(min_row=4, max_row=None, min_col=5, max_col=None):
        for cell in rows:
            cell.alignment = Alignment(wrapText=True)

    # Apply Table Formatting
    tab = Table(displayName="DiagnosticsTable", ref=f"A4:E54")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    excel_file_name = f"out/reports/{ILLEGAL_CHARACTERS_RE.sub(r'', SOC_SoftwareVersion)}_EOL_Parameters_for_MPI_{timestamp}.xlsx"
    wb.save(excel_file_name)

    return True

# ************************************************************* TEST CASES #01 *************************************************
@pytest.mark.hpcc_mpi
@pytest.mark.order(1)
@pytest.mark.parametrize("param_id, EOL_Parameter, ExpDataLength, ActDataLength, Data_StartPosition, Enter_ManDiag, read_data_cmd", TEST_CASES, ids=custom_ids)
def test_EOL_Parameter_read(ManDiag_component, param_id, EOL_Parameter, ExpDataLength, ActDataLength, Data_StartPosition, Enter_ManDiag, read_data_cmd):
    """ VIP: Read EOL parameter data """

    global index_row, wb, ws, SOC_SoftwareVersion, timestamp
    if param_id == 1:
        Create_excelFile_with_table_headings(ManDiag_component)

    ws.cell(row=index_row, column=1, value=param_id)
    ws.cell(row=index_row, column=2, value=EOL_Parameter)
    ws.cell(row=index_row, column=3, value=ActDataLength)
    ws.cell(row=index_row, column=4, value=read_data_cmd)

    response_data = ManDiag_component.send_request_command_and_get_response(read_data_cmd, retries=4)
    if response_data["DataLength"] == ExpDataLength:
        hex_str = response_data["Databytes"]
        ws.cell(row=index_row, column=5, value=hex_str[Data_StartPosition:])
        app_logger.info(f"Data written to Excel {hex_str[Data_StartPosition:]}")
    else:
        assert False, f"Failed to read '{EOL_Parameter}' data"

    index_row += 1

    if param_id == 50:
        assert apply_tabularFormate_alignment_ToExcelData(ManDiag_component), f"error in saving excel file data"
