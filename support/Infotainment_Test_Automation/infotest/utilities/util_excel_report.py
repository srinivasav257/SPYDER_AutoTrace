"""
=====================================================================================================================
File Name   : excel_report.py
Description : This File contains ExcelReport class - This module will handles Creating Excel reports for test results.
Author      : Srinivasa V
Created On  : 2025-01-01

Copyright (C) 2025 APTIV - This file is part of the MnM HPCC project - Manufacture Diagnostics Component.
Manufacture Diagnostics Project is free software: you can redistribute it and/or modify it under the terms of Diagnostics
Department ASUX.

======================================================================================================================
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
from infotest.core.trace_logger import app_logger

class ExcelReport:
    """ Class for generating and managing Excel test reports. """

    def __init__(self, file_name="Test_Report.xlsx"):
        self.file_name = file_name
        self.wb = Workbook()
        self.summary_sheet = self.wb.active
        self.summary_sheet.title = "Test Summary"
        self.setup_summary_sheet()

    def setup_summary_sheet(self):
        """ Creates the initial structure for the summary sheet. """
        self.summary_sheet.append(["Software Version", "Total Tests", "Passed", "Failed", "No Response"])
        self.apply_header_style(self.summary_sheet, "A1:E1")

    def apply_header_style(self, ws, cell_range):
        """ Applies styling to headers. """
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for row in ws[cell_range]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def add_test_result(self, software_version, total_tests, passed, failed, no_response):
        """ Adds a new row of test results to the summary sheet. """
        self.summary_sheet.append([software_version, total_tests, passed, failed, no_response])
        self.save_report()

    def save_report(self):
        """ Saves the Excel report to disk. """
        try:
            self.wb.save(self.file_name)
            app_logger.info(f"Excel report saved: {self.file_name}")
        except Exception as e:
            app_logger.error(f"Failed to save Excel report: {e}")

    def load_existing_report(self, file_path):
        """ Loads an existing Excel report. """
        if os.path.exists(file_path):
            self.wb = load_workbook(file_path)
            self.summary_sheet = self.wb.active
            app_logger.info(f"Loaded existing report: {file_path}")
        else:
            app_logger.warning(f"File not found: {file_path}. Creating a new report.")

    def close_report(self):
        """ Closes the workbook to free resources. """
        self.wb.close()

# Adjust column widths based on content
def adjust_column_row_width_as_per_length(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    return None

# Apply blue color formate
def apply_blue_color_formate(ws, cell):
    # Border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    ws[cell].alignment = Alignment(horizontal='center')
    ws[cell].font = Font(bold=True, size=14, color="f7f7f7")
    ws[cell].fill = PatternFill("solid", fgColor="4F81BD")  # Blue header background
    ws[cell].border = thin_border

"""
Revision History:
|---------------------------------------------------------------------------------------|
| Version | Date       | Author            | Changes                                    |
|---------|------------|-------------------|--------------------------------------------|
| 1.0     | 2025-01-01 | Srinivas V        | Initial release                            |

"""
